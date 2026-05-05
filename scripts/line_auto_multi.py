"""
LINE 多客戶自動回覆腳本（SOP 話術引擎）
用法：python line_auto_multi.py <監控時間HH:MM> [SOP設定檔.json]
範例：python line_auto_multi.py 23:30 scripts/line_sop/織夢小棧.json

流程：
1. 開啟聊天頁
2. 截圖偵測有綠色未讀標記的對話
3. 點第一個綠色標記 → 進入聊天
4. OCR 讀取對話內容 → 判斷 SOP 步驟
5. Claude AI 生成回覆 → 發送
6. 回到聊天列表
7. 點下一個綠色標記 → 重複 3-6
8. 全部處理完 → 等待新的未讀
9. 持續循環直到監控時間結束
"""
import sys
import io
import os
import re
import json
import time
import signal
import ctypes
import hashlib
import types
import atexit
import gc

# DPI + 編碼
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(0)
except Exception:
    pass

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import win32gui
import pyautogui
import pyperclip
from datetime import datetime
from dotenv import load_dotenv
from pathlib import Path
import anthropic

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(SCRIPT_DIR.parent))

load_dotenv(Path("C:/Users/blue_/claude-telegram-bot/.env"))
client = anthropic.Anthropic()
TMPDIR = "C:/Users/blue_/Desktop/測試檔案"
DEFAULT_SOP = "C:/Users/blue_/claude-telegram-bot/scripts/line_sop/織夢小棧.json"
EXCEL_PATH = r"C:\Users\blue_\Desktop\客戶資料.xlsx"

# ============================================================
# Excel 讀寫（客戶資料）
# ============================================================
import openpyxl

def write_customer_to_excel(name, birthday, phone, area, line_name, ad_type=""):
    """寫入客戶資料到 Excel（編號先空），回傳行號。ad_type=廣告種類（SOP 名稱，區分客戶從哪個帳號來）"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    row_num = ws.max_row + 1
    ws.cell(row=row_num, column=1, value="")           # 編號（先空）
    ws.cell(row=row_num, column=2, value=name)          # 姓名
    ws.cell(row=row_num, column=3, value=birthday)      # 出生月日
    ws.cell(row=row_num, column=4, value=phone)         # 電話
    ws.cell(row=row_num, column=5, value=area)          # 地區
    ws.cell(row=row_num, column=6, value=line_name)     # LINE名稱
    ws.cell(row=row_num, column=7, value=datetime.now().strftime("%Y-%m-%d"))  # 報名日期
    ws.cell(row=row_num, column=8, value=ad_type)       # 廣告種類（SOP 名稱）
    ws.cell(row=row_num, column=9, value="")            # 撞腳（先空，撞號比對後填 Y/N）
    wb.save(EXCEL_PATH)
    return row_num

def finalize_customer_normal(row_num, customer_id):
    """正常客戶：col 1=電話後五碼、col 9=N"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row=row_num, column=1, value=customer_id)
    ws.cell(row=row_num, column=9, value="N")
    wb.save(EXCEL_PATH)

def finalize_customer_collision(row_num):
    """撞號客戶：col 1 留空、col 9=Y（被撞那筆 A 不動）"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row=row_num, column=9, value="Y")
    wb.save(EXCEL_PATH)

def check_duplicate_phone_in_excel(phone, exclude_row=None):
    """用 col 4 全 10 碼比對 Excel，回傳被撞那筆 row（int）或 None。"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if exclude_row is not None and row == exclude_row:
            continue
        cell = ws.cell(row=row, column=4).value
        if cell and str(cell).strip() == phone:
            wb.close()
            return row
    wb.close()
    return None

# ============================================================
# 層 3：提取資料驗證器（hybrid validator）
# ============================================================
def validate_extraction(extracted, them_texts=None):
    """驗證 Step 2 提取結果是否合法。回傳錯誤列表（空 = OK）。

    them_texts: 客戶訊息列表（給層 5「值必須在原文」用）

    擋下 AI 常見的 hallucination：
    - 生日月日不合法（如 09/87）
    - 電話格式錯
    - 生日跟電話前 4 碼重複（AI 容易把電話前 4 碼當生日）
    - 🆕 層 5：生日 4 碼必須真的出現在客戶訊息原文中（防 AI 拼湊虛構）
    - 🆕 層 5：電話 10 碼必須真的出現在客戶訊息原文中
    """
    errors = []

    # 1. 生日合法性（必須是 MM/DD，月 1-12，日 1-31）
    birthday = str(extracted.get("birthday", "")).strip()
    bday_4 = ""  # 4 碼純數字版（拿來跟電話前 4 碼比對）
    try:
        m_str, d_str = birthday.split("/")
        m, d = int(m_str), int(d_str)
        if not (1 <= m <= 12 and 1 <= d <= 31):
            errors.append(f"生日 {birthday} 月日不合法（月必須 1-12，日 1-31）")
        else:
            bday_4 = f"{m:02d}{d:02d}"
    except (ValueError, AttributeError):
        errors.append(f"生日 '{birthday}' 不是 MM/DD 格式")

    # 2. 電話合法性（09 開頭 10 碼純數字）
    phone = str(extracted.get("phone", "")).strip()
    phone_digits = "".join(c for c in phone if c.isdigit())
    if not phone_digits.startswith("09") or len(phone_digits) != 10:
        errors.append(f"電話 '{phone}' 不是 09 開頭 10 碼")

    # 3. ⭐ 生日不能跟電話前 4 碼重複（AI 把電話前 4 碼誤當生日的常見錯誤）
    if bday_4 and phone_digits.startswith(bday_4):
        errors.append(f"生日 {bday_4} 跟電話前 4 碼相同 → AI 可能把電話前 4 碼誤當生日")

    # 4. ⭐⭐ 層 5：值必須真的出現在客戶訊息原文中（純數字 OR 帶分隔符多種寫法）
    if them_texts:
        all_them_text = " ".join(them_texts)
        all_them_digits = "".join(c for c in all_them_text if c.isdigit())

        # 4a. 生日：嘗試多種寫法（4/6、5/20、80/11/19、4月6日 都相容）
        if bday_4:
            try:
                bm = int(m_str)
                bd = int(d_str)
                # 純數字版（4 碼補零標準）
                digit_candidate = f"{bm:02d}{bd:02d}"   # 0406, 0520, 1119
                # 原文版（含各種分隔符 / 月日 - .）
                text_candidates = [
                    f"{bm:02d}/{bd:02d}", f"{bm}/{bd:02d}",
                    f"{bm:02d}/{bd}",     f"{bm}/{bd}",
                    f"{bm:02d}月{bd:02d}日", f"{bm}月{bd}日",
                    f"{bm:02d}月{bd}日",     f"{bm}月{bd:02d}日",
                    f"{bm:02d}-{bd:02d}", f"{bm}-{bd}",
                    f"{bm:02d}-{bd}",     f"{bm}-{bd:02d}",
                    # 「.」分隔符（民國年常見寫法 45.5.19、85.07.24 等）
                    f"{bm:02d}.{bd:02d}", f"{bm}.{bd:02d}",
                    f"{bm:02d}.{bd}",     f"{bm}.{bd}",
                ]
                found_in_digits = digit_candidate in all_them_digits
                found_in_text = any(c in all_them_text for c in text_candidates)
                if not (found_in_digits or found_in_text):
                    errors.append(
                        f"生日 {birthday} 沒出現在客戶訊息中（嘗試純數字 {digit_candidate} "
                        f"和分隔符寫法都找不到），AI 可能虛構"
                    )
            except (ValueError, NameError):
                pass  # step 1 已標記非法格式

        # 4b. 電話 10 碼（電話格式較單一，純數字比對已穩）
        if phone_digits and len(phone_digits) == 10 and phone_digits not in all_them_digits:
            errors.append(
                f"電話 {phone_digits} 沒出現在客戶訊息純數字中，AI 可能虛構"
            )

    return errors

# ============================================================
# 層 4：程式偵測「資料是否齊全」(LLM fallback)
# ============================================================
# 台灣 22 個縣市（含金門連江）
_TW_AREA_KEYWORDS = [
    "台北", "臺北", "新北", "桃園", "台中", "臺中", "台南", "臺南", "高雄",
    "新竹", "苗栗", "彰化", "南投", "雲林", "嘉義", "屏東", "宜蘭", "花蓮",
    "台東", "臺東", "基隆", "金門", "連江", "澎湖", "馬祖",
]

def has_complete_data(new_them, history):
    """偵測 new_them + history 對方訊息中是否含完整四項（姓名+生日+電話+地區）。

    用途：AI 漏出 [DATA_READY] 時，程式 fallback 兜底。
    """
    them_texts = list(new_them) + [m["text"] for m in history if m.get("sender") == "them"]
    all_text = " ".join(them_texts)

    # 1. 09 開頭 10 碼電話（容許空格、橫線）
    has_phone = bool(re.search(r'09\d{2}[\s\-]?\d{3}[\s\-]?\d{3}', all_text)) \
                or bool(re.search(r'09\d{8}', "".join(c for c in all_text if c.isdigit() or c in " -")))

    # 2. 4 碼合法生日（MMDD 格式：月 01-12 + 日 01-31，且前後不黏其他數字）
    has_birthday = bool(re.search(
        r'(?<![\d])(0[1-9]|1[0-2])(0[1-9]|[12][0-9]|3[01])(?![\d])',
        all_text
    ))

    # 3. 中文姓名（連續 2-4 個中文字）
    has_name = bool(re.search(r'[一-鿿]{2,4}', all_text))

    # 4. 地區關鍵字（台灣縣市）
    has_area = any(kw in all_text for kw in _TW_AREA_KEYWORDS)

    return has_phone and has_birthday and has_name and has_area

def read_customer_from_excel(customer_id):
    """用編號讀取客戶資料"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in range(ws.max_row, 1, -1):
        if str(ws.cell(row=row, column=1).value) == str(customer_id):
            return {
                "name": ws.cell(row=row, column=2).value,
                "birthday": ws.cell(row=row, column=3).value,
                "phone": ws.cell(row=row, column=4).value,
                "area": ws.cell(row=row, column=5).value,
                "line_name": ws.cell(row=row, column=6).value,
            }
    return None

# ============================================================
# GPU 記憶體清理
# ============================================================
def _cleanup_gpu():
    try:
        gc.collect()
        import paddle
        paddle.device.cuda.empty_cache()
        print("[Cleanup] GPU 記憶體已釋放", flush=True)
    except Exception:
        pass

atexit.register(_cleanup_gpu)

# ============================================================
# 優雅停止機制
# ============================================================
STOP_FILE = "C:/Users/blue_/Desktop/測試檔案/.stop_line_multi"
_should_stop = False

def _signal_handler(signum, frame):
    global _should_stop
    _should_stop = True
    print(f"\n[STOP] 收到信號 {signum}，準備停止...", flush=True)
    _cleanup_gpu()

signal.signal(signal.SIGINT, _signal_handler)
if hasattr(signal, "SIGBREAK"):
    signal.signal(signal.SIGBREAK, _signal_handler)

def should_stop():
    global _should_stop
    if _should_stop:
        return True
    if os.path.exists(STOP_FILE):
        try:
            os.remove(STOP_FILE)
        except OSError:
            pass
        _should_stop = True
        print("[STOP] 偵測到停止旗標檔案，準備停止...", flush=True)
        return True
    return False

# ============================================================
# SOP 設定檔（從 line_auto_chat.py 引用）
# ============================================================
from line_auto_chat import (
    load_sop, build_system_prompt, generate_reply,
    filter_reply, send_reply, send_multi_reply, send_image,
    grab_chat_area, ocr_extract_messages, chat_hash,
)

# ============================================================
# 時間判斷
# ============================================================
def time_to_minutes(t_str):
    h, m = map(int, t_str.split(":"))
    return h * 60 + m

def is_before_stop_time(stop_time):
    now_min = time_to_minutes(datetime.now().strftime("%H:%M"))
    stop_min = time_to_minutes(stop_time)
    return now_min < stop_min

# ============================================================
# 偵測聊天列表中有未讀標記的對話
# ============================================================
def find_unread_conversations(monitor=None):
    """
    切到聊天頁，用 line_locate.py 的 find_unread_badges 偵測綠色未讀標記。

    回傳：(regions, unread_list)
        regions: 當前定位
        unread_list: [{"y": int, "center": (x, y)}] 有未讀的對話座標
    """
    from line_locate import (
        locate_line_regions, switch_page, find_unread_badges,
    )

    regions = locate_line_regions(monitor)

    # 確保在聊天頁
    if regions["current_page"] != "chat":
        regions = switch_page(regions, "chat", monitor)

    # 用定位腳本的像素偵測找綠色徽章（LINE16 方式）
    unread_list = find_unread_badges(monitor)

    return regions, unread_list


# ============================================================
# 處理單一客戶的對話
# ============================================================
def handle_one_customer(conv, regions, system_prompt, sop, all_histories, monitor=None):
    """
    點進一個客戶的對話，讀取內容，判斷 SOP 步驟，回覆。

    參數：
        conv: {"name", "unread", "center"} 從聊天列表取得
        regions: 當前的 regions
        system_prompt: SOP system prompt
        sop: SOP JSON
        all_histories: {客戶名: [conversation_history]} 所有客戶的對話歷史
        monitor: 螢幕編號

    回傳：更新後的 regions
    """
    from line_locate import locate_line_regions, ocr_scan_panel, screenshot_line

    cx, cy = conv["center"]
    print(f"\n[Customer] 點擊未讀對話 at ({cx}, {cy})", flush=True)

    # Step 1: 點擊該對話進入聊天
    pyautogui.click(cx, cy)
    time.sleep(1.5)

    # Step 2: 重新定位（進入聊天後）
    regions = locate_line_regions(monitor)

    # 從 chat_title 區域讀取客戶名稱
    ct = regions.get("chat_title", {})
    full_img, line_crop, (il, it, ir, ib), mon = screenshot_line(monitor)
    sx_r = full_img.size[0] / mon["width"]
    sy_r = full_img.size[1] / mon["height"]
    ct_il = max(0, int((ct.get("left", 0) - mon["left"]) * sx_r) - il)
    ct_it = max(0, int((ct.get("top", 0) - mon["top"]) * sy_r) - it)
    ct_ir = min(line_crop.size[0], int((ct.get("right", 0) - mon["left"]) * sx_r) - il)
    ct_ib = min(line_crop.size[1], int((ct.get("bottom", 0) - mon["top"]) * sy_r) - it)
    if ct_ir > ct_il and ct_ib > ct_it:
        title_crop = line_crop.crop((ct_il, ct_it, ct_ir, ct_ib))
        title_items = ocr_scan_panel(title_crop)
        name = " ".join(item["text"] for item in title_items).strip() if title_items else "unknown"
    else:
        name = "unknown"
    print(f"[Customer] 客戶名稱: {name}", flush=True)

    # 群組過濾：黑名單（部分匹配） + (數字) 特徵偵測
    SKIP_GROUPS = ["友資群", "友资群", "友資", "友资", "好朋友的群組", "好朋友的群组"]
    if any(g in name or name in g for g in SKIP_GROUPS) or re.search(r"\(\d+\)", name):
        print(f"[Customer] {name} 是群組，跳過", flush=True)
        return regions

    # 撞腳客戶過濾：已被標記為「撞腳」的客戶 → 完全不處理（避免重複發敷衍訊息）
    if name.startswith("撞腳") or name.startswith("撞脚"):
        print(f"[Customer] {name} 已標記為撞腳客戶，跳過不處理", flush=True)
        return regions

    # Step 3: 截圖對話區 + OCR 讀取內容
    chat_img = grab_chat_area(regions, monitor)
    current_messages = ocr_extract_messages(chat_img)
    print(f"[Customer] OCR 讀到 {len(current_messages)} 條訊息", flush=True)

    # 偵測是否為好友（非好友會有 LINE 警告框）
    NOT_FRIEND_KEYWORDS = ["加入好友", "請先確認", "封鎖", "檢舉", "詐騙行為", "用戶可疑"]
    all_text = " ".join(m["text"] for m in current_messages)
    is_friend = not any(kw in all_text for kw in NOT_FRIEND_KEYWORDS)
    print(f"[Customer] 好友狀態: {'是好友' if is_friend else '不是好友'}", flush=True)

    # Step 4: 取得或建立該客戶的對話歷史
    if name not in all_histories:
        all_histories[name] = []
    history = all_histories[name]

    # Step 5: 找出新的對方訊息
    # 如果歷史是空的，用 OCR 讀到的所有對方訊息
    if not history:
        new_them = [m["text"] for m in current_messages if m["sender"] == "them"]
        # 把所有訊息加入歷史
        for m in current_messages:
            history.append(m)
    else:
        # 用最後一條已知訊息定位新訊息
        from difflib import SequenceMatcher
        last_known = history[-1]
        last_text = last_known["text"]

        match_idx = -1
        for i in range(len(current_messages) - 1, -1, -1):
            ratio = SequenceMatcher(None, current_messages[i]["text"], last_text).ratio()
            if ratio > 0.6:
                match_idx = i
                break

        if match_idx >= 0 and match_idx < len(current_messages) - 1:
            new_msgs = current_messages[match_idx + 1:]
            new_them = [m["text"] for m in new_msgs if m["sender"] == "them"]
            for m in new_msgs:
                history.append(m)
        else:
            # 找不到匹配，用所有對方訊息
            new_them = [m["text"] for m in current_messages if m["sender"] == "them"]

    if not new_them:
        print(f"[Customer] 沒有新的對方訊息，跳過", flush=True)
        return regions

    print(f"[Customer] 新訊息: {new_them}", flush=True)

    # 純貼圖 → 用 Haiku 4.5 Vision 解讀貼圖含意，注入到訊息讓 AI 看得懂
    # （客戶傳文字不觸發 Vision，省 token）
    from line_auto_chat import is_only_sticker, analyze_sticker, send_multi_reply
    if is_only_sticker(new_them):
        meaning = analyze_sticker(regions, monitor)
        print(f"[Customer] 純貼圖 → Vision 解讀為「{meaning}」", flush=True)
        new_them = [f"[貼圖含意：{meaning}]"]

    # Step 6: Claude AI 生成回覆
    reply = generate_reply(system_prompt, history, new_them)

    if not reply or len(reply) <= 1:
        time.sleep(0.5)
        return regions

    # 偵測 [END] 標記
    if "[END]" in reply:
        print(f"[Customer] SOP 結束（{name}）", flush=True)
        return regions

    # ============================================================
    # 偵測 [DATA_READY] 暗號 → 走「先寫 Excel → 比對撞號 → 給編號」流程
    # 雙保險：AI 出 [DATA_READY] OR 程式偵測四項齊全（fallback 兜底 AI 漏判）
    # ============================================================
    ai_signal = "[DATA_READY]" in reply
    fallback_trigger = (not ai_signal) and has_complete_data(new_them, history)
    if ai_signal or fallback_trigger:
        if ai_signal:
            print(f"[Customer] AI 判定資料齊全（[DATA_READY]）→ 進入報名流程", flush=True)
        else:
            print(f"[Customer] ⚠️ AI 漏出 [DATA_READY] 但程式偵測四項齊全 → fallback 觸發報名流程", flush=True)
            print(f"[Customer] AI 原始 reply: {reply[:120]}", flush=True)
        # ⚠️ [DATA_READY] / 一般 reply 不送給客戶、不寫進 history（避免污染歷史）

        from line_locate import (
            share_contact_card, switch_page, locate_line_regions,
            rename_friend, find_add_friend_btn,
        )

        # === Step 2: Claude API 結構化輸出提取報名資料（schema 強化 + few-shot + validator retry）===
        history_lines = []
        for msg in history[-20:]:
            label = "[客戶]" if msg["sender"] == "them" else "[小編]"
            history_lines.append(f"{label} {msg['text']}")
        history_text = "\n".join(history_lines)
        new_text = "\n".join(f"• {m}" for m in new_them)
        print(f"[Customer] Step2: 新訊息數={len(new_them)}", flush=True)

        # 層 1：強化 schema（pattern 擋非法格式）
        extraction_schema = {
            "type": "object",
            "properties": {
                "customer_name": {
                    "type": "string",
                    "description": "客戶姓名（2-4 個中文字）"
                },
                "birthday": {
                    "type": "string",
                    "pattern": "^(0[1-9]|1[0-2])/(0[1-9]|[12][0-9]|3[01])$",
                    "description": "出生月日 MM/DD（月 01-12，日 01-31）。⚠️ 絕對不能拿電話的前 4 碼當生日，必須是另一個獨立的 4 碼數字。"
                },
                "phone": {
                    "type": "string",
                    "pattern": "^09\\d{8}$",
                    "description": "手機號碼，09 開頭 10 碼純數字（不要含空格、橫線）"
                },
                "area": {
                    "type": "string",
                    "description": "客戶想參加的地點（縣市名稱）"
                }
            },
            "required": ["customer_name", "birthday", "phone", "area"],
            "additionalProperties": False
        }

        # 層 2：few-shot 範例（含邊緣情境）
        base_prompt = (
            f"以下是對話紀錄：\n\n{history_text}\n\n"
            f"客戶剛發了新訊息:\n{new_text}\n\n"
            f"從客戶的訊息中提取報名資料。\n\n"
            f"拆分規則:\n"
            f"- 中文 = 姓名（2-4 個中文字）\n"
            f"- 4碼數字 = 生日月日（格式 MM/DD，如 0910 → 09/10）\n"
            f"- 09開頭10碼 = 手機號碼\n"
            f"- 地點相關中文 = 地區（縣市）\n"
            f"- 數字可能黏在一起，自己拆分\n"
            f"- 忽略課程圖片和系統訊息，只看客戶自己打的內容\n\n"
            f"範例（包含邊緣情境）:\n"
            f"✅ 王小明 0915 0980588129 → name=王小明, birthday=09/15, phone=0980588129\n"
            f"✅ 陳采汝 620926 0968903365 → name=陳采汝, birthday=09/26, phone=0968903365\n"
            f"⚠️ 王大路 0912 0987652714 → name=王大路, birthday=09/12, phone=0987652714\n"
            f"   注意：09/12 是生日，不是 09/87！0987 是電話前 4 碼，不是生日。\n"
            f"⚠️ 數字黏一起 09120987652714 → birthday=09/12, phone=0987652714（前 4 碼=生日，後 10 碼=電話）\n"
            f"⚠️ 電話前 4 碼（如 0987、0915、0912）永遠不能當生日，生日必須是『獨立的另一個 4 碼數字』"
        )

        # 層 3：retry + validator
        extracted = None
        last_errors = []
        for attempt in range(3):
            try:
                feedback = ""
                if last_errors:
                    feedback = f"\n\n⚠️ 你上次提取的錯誤：{'; '.join(last_errors)}\n請仔細重新拆分，避免同樣錯誤。"

                extract_r = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=200,
                    messages=[{"role": "user", "content": base_prompt + feedback}],
                    output_config={
                        "format": {
                            "type": "json_schema",
                            "schema": extraction_schema
                        }
                    }
                )
                candidate = json.loads(extract_r.content[0].text)
                print(f"[Customer] Step2: AI 提取（第{attempt+1}次）: {candidate}", flush=True)

                # 程式 validator 雙重把關
                # 層 5：把客戶訊息（new_them + history 對方訊息）傳進去做「值必須在原文」檢查
                # dict.fromkeys 保留順序去重（new_them 跟 history 可能有同樣訊息，避免重複收錄）
                step2_them_texts = list(dict.fromkeys(
                    list(new_them) + [m["text"] for m in history if m.get("sender") == "them"]
                ))
                errors = validate_extraction(candidate, them_texts=step2_them_texts)
                if not errors:
                    extracted = candidate
                    print(f"[Customer] Step2: 提取通過 validator ✅", flush=True)
                    break
                else:
                    last_errors = errors
                    print(f"[Customer] Step2: validator 拒絕（第{attempt+1}次）: {errors}", flush=True)
                    if attempt < 2:
                        time.sleep(1)
            except Exception as e:
                last_errors = [f"上次 API 錯誤: {str(e)[:100]}"]
                print(f"[Customer] Step2: 提取失敗（第{attempt+1}次）: {e}", flush=True)
                if attempt < 2:
                    time.sleep(2)

        if not extracted:
            print(f"[Customer] Step2: 3 次重試後仍失敗（最後錯誤：{last_errors}），放棄報名流程", flush=True)
            return regions

        # 取乾淨的 phone（schema pattern 已限制 09 開頭 10 碼，這裡保險再清一次）
        phone = "".join(c for c in str(extracted.get("phone", "")) if c.isdigit())
        print(f"[Customer] Step2: 最終 → name={extracted.get('customer_name')}, birthday={extracted.get('birthday')}, phone={phone}, area={extracted.get('area')}", flush=True)

        # === Step 3: 寫 Excel（col 1+9 留空，待撞號比對後補）===
        ad_type = sop.get("name", "")
        excel_row = write_customer_to_excel(
            name=extracted.get("customer_name", ""),
            birthday=extracted.get("birthday", ""),
            phone=phone,
            area=extracted.get("area", ""),
            line_name=name,
            ad_type=ad_type,
        )
        print(f"[Customer] Step3: 已寫入 Excel 第{excel_row}行（col 1+9 待補）", flush=True)

        # === Step 4: 用 col 4 全 10 碼比對撞號（排除自己這筆）===
        duplicate_row = check_duplicate_phone_in_excel(phone, exclude_row=excel_row)

        if duplicate_row:
            # === 撞號分支 ===
            print(f"[Customer] Step4: 撞號（電話 {phone} 跟第 {duplicate_row} 行重複）→ 標撞腳", flush=True)
            finalize_customer_collision(excel_row)

            # 發敷衍訊息
            collision_msg = "您的資料我們已收到，工作人員會盡快與您聯繫💕"
            send_multi_reply(collision_msg, regions)
            history.append({"text": collision_msg, "sender": "me", "y": 0})
            time.sleep(1)

            # 非好友 → 加好友
            if not is_friend:
                pos = find_add_friend_btn(regions, monitor)
                if pos:
                    pyautogui.click(*pos)
                    print(f"[Customer] 撞號客戶非好友 → 點加好友 at {pos}", flush=True)
                    time.sleep(7)
                else:
                    print(f"[Customer] 撞號客戶非好友但找不到加好友按鈕", flush=True)

            # 改名「撞腳 MM-DD」
            rename_name = f"撞腳 {datetime.now().strftime('%m-%d')}"
            regions = locate_line_regions(monitor)
            rename_ok = rename_friend(regions, rename_name, monitor, current_name=name)
            if rename_ok:
                print(f"[Customer] 已改名為「{rename_name}」", flush=True)
            else:
                print(f"[Customer] ⚠️ 改名「{rename_name}」失敗 → 請手動改名", flush=True)

            return regions

        # === 不撞號分支 ===
        customer_id = phone[-5:]
        finalize_customer_normal(excel_row, customer_id)
        print(f"[Customer] Step4: 不撞號 → 編號 = {customer_id}（電話後五碼）→ Excel col 1={customer_id}, col 9=N", flush=True)

        # === Step 5: 從 SOP json 拿 give_id 範本送出 ===
        give_id_step = next((s for s in sop.get("steps", []) if s.get("id") == "give_id"), None)
        if give_id_step:
            for tpl in give_id_step.get("replies", []):
                msg = tpl.replace("{id}", customer_id)
                send_multi_reply(msg, regions)
                history.append({"text": msg, "sender": "me", "y": 0})
                time.sleep(0.5)
            print(f"[Customer] Step5: 已送出 give_id 範本（編號 {customer_id}）", flush=True)
        else:
            # fallback: SOP 沒 give_id step
            id_kw = sop.get("rules", {}).get("id_keyword", "編號")
            fb_msg = f"課程名額成功登記✅\n\n這是您的{id_kw}：{customer_id}\n麻煩請複製傳給溫妮"
            send_multi_reply(fb_msg, regions)
            history.append({"text": fb_msg, "sender": "me", "y": 0})
            print(f"[Customer] Step5: SOP 沒 give_id step，用 fallback 訊息（編號 {customer_id}）", flush=True)

        # === Step 6: 非好友 → 加好友 ===
        if not is_friend:
            pos = find_add_friend_btn(regions, monitor)
            if pos:
                pyautogui.click(*pos)
                print(f"[Customer] Step6: 點擊「加入好友」at {pos}", flush=True)
                time.sleep(3)
            else:
                print(f"[Customer] Step6: 找不到「加入好友」按鈕，跳過加好友", flush=True)

        # === Step 7: rename_friend（把 LINE 名稱改成 編號+日期）===
        rename_name = f"{customer_id} {datetime.now().strftime('%m-%d')}"
        regions = locate_line_regions(monitor)
        rename_friend(regions, rename_name, monitor, current_name=name)
        print(f"[Customer] Step7: 已改名為 {rename_name}", flush=True)
        time.sleep(1)
        search_name = rename_name

        # === Step 8: 分享溫妮給客戶 ===
        regions = locate_line_regions(monitor)
        result = share_contact_card(regions, "溫妮", search_name, monitor)
        if result:
            print(f"[Customer] Step8: 已分享溫妮給 {search_name}", flush=True)
        else:
            print(f"[Customer] Step8: 分享溫妮給 {search_name} 失敗", flush=True)

        # === Step 9: 讀 Excel → 組報名資訊 → 發友資群 ===
        cust_data = read_customer_from_excel(customer_id)
        if cust_data:
            info = (
                f"【新報名】\n"
                f"✏ 姓名：{cust_data['name']}\n"
                f"✏ 出生月/日：{cust_data['birthday']}\n"
                f"✏ 聯絡電話：{cust_data['phone']}\n"
                f"✏ 想參加的地點：{cust_data['area']}\n"
                f"✏ 編號：{customer_id}"
            )
        else:
            info = f"【新報名】{name}\n編號：{customer_id}"
            print(f"[Customer] Step9: Excel 找不到編號 {customer_id}，使用簡易格式", flush=True)

        regions = locate_line_regions(monitor)
        if regions["current_page"] != "chat":
            regions = switch_page(regions, "chat", monitor)

        # 點友資群（聊天頁第一個對話）
        lp = regions["left_panel"]
        first_item_x = (lp["left"] + lp["right"]) // 2
        first_item_y = lp["top"] + 29
        pyautogui.click(first_item_x, first_item_y)
        time.sleep(1.5)
        regions = locate_line_regions(monitor)

        send_reply(info, regions)
        print(f"[Customer] Step9: 已發送報名資訊到友資群", flush=True)
        time.sleep(1)

        # === Step 10: 分享客戶好友資訊到友資群 ===
        regions = switch_page(regions, "friend", monitor)
        result = share_contact_card(regions, search_name, "友資群", monitor)
        if result:
            print(f"[Customer] Step10: 已分享 {search_name} 的好友資訊到友資群", flush=True)
        else:
            print(f"[Customer] Step10: 分享 {search_name} 的好友資訊到友資群失敗", flush=True)

        # === Step 11: 回到聊天頁 ===
        regions = locate_line_regions(monitor)
        if regions["current_page"] != "chat":
            regions = switch_page(regions, "chat", monitor)
        print(f"[Customer] Step11: 已回到聊天頁", flush=True)

        time.sleep(0.5)
        return regions

    # ============================================================
    # 一般聊天分支：AI 沒出 [DATA_READY] → 照舊送 reply
    # ============================================================
    reply = filter_reply(reply)
    if not reply:
        print(f"[Customer] 過濾後為空，跳過", flush=True)
        return regions

    print(f"[Customer] 回覆: {reply[:80]}", flush=True)
    send_multi_reply(reply, regions)

    # 歡迎詞 → 發課程圖片
    if "編織" in reply or "歡迎" in reply:
        img_path = sop.get("course_info", {}).get("image", "")
        if img_path:
            if not os.path.isabs(img_path):
                img_path = str(SCRIPT_DIR.parent / img_path)
            send_image(img_path, regions)
            print(f"[Customer] 已發送課程圖片: {img_path}", flush=True)

    # 回覆加入歷史
    for part in reply.split("|||"):
        part = part.strip()
        if part and not (part.startswith("{") and part.endswith("}")):
            history.append({"text": part, "sender": "me", "y": 0})

    time.sleep(0.5)
    return regions
    return regions


# ============================================================
# 主流程
# ============================================================
def main(stop_time, sop_path=DEFAULT_SOP, monitor=None):
    from line_locate import (
        locate_line_regions, switch_page, find_line_window,
        screenshot_line_window,
    )
    import win32con

    # 清殘留旗標
    if os.path.exists(STOP_FILE):
        try:
            os.remove(STOP_FILE)
        except OSError:
            pass

    print("=" * 50, flush=True)
    print(f"LINE 多客戶自動回覆", flush=True)
    print(f"監控到：{stop_time}", flush=True)
    print(f"SOP：{sop_path}", flush=True)
    print("=" * 50, flush=True)

    # 多開 LINE 設定：每個 box 對應的 SOP 檔案路徑
    BOX_SOP = {
        "blue_1": "C:/Users/blue_/claude-telegram-bot/scripts/line_sop/Resin藝術學堂.json",
        "blue_2": "C:/Users/blue_/claude-telegram-bot/scripts/line_sop/織夢小棧.json",
    }

    # 預載所有 box 的 SOP 跟 system_prompt（避免每輪迴圈重複載入）
    box_sop_cache = {}      # box -> sop dict
    box_prompt_cache = {}   # box -> system_prompt str
    for _box, _sop_path in BOX_SOP.items():
        try:
            _sop = load_sop(_sop_path)
            box_sop_cache[_box] = _sop
            box_prompt_cache[_box] = build_system_prompt(_sop)
            print(f"[Init] {_box} → SOP: {_sop['name']}（{len(_sop['steps'])} 步驟）", flush=True)
        except Exception as e:
            print(f"[Init] {_box} 載入 SOP 失敗: {e}", flush=True)

    # 預設用第一個 box 的 SOP（給單開模式向下相容）
    sop = next(iter(box_sop_cache.values()))
    system_prompt = next(iter(box_prompt_cache.values()))

    # 置前 LINE
    line = find_line_window()
    if not line:
        print("[ERROR] 找不到 LINE 視窗", flush=True)
        return False

    SWP_FLAGS = win32con.SWP_NOMOVE | win32con.SWP_NOSIZE
    try:
        win32gui.SetWindowPos(line[0], win32con.HWND_NOTOPMOST, 0, 0, 0, 0, SWP_FLAGS)
        win32gui.SetWindowPos(line[0], win32con.HWND_TOPMOST, 0, 0, 0, 0, SWP_FLAGS)
        win32gui.SetWindowPos(line[0], win32con.HWND_NOTOPMOST, 0, 0, 0, 0,
                              SWP_FLAGS | win32con.SWP_SHOWWINDOW)
        win32gui.SetForegroundWindow(line[0])
    except Exception:
        pass
    time.sleep(0.5)

    # 所有客戶的對話歷史
    all_histories = {}

    # 主迴圈
    POLL_INTERVAL = 10  # 每 10 秒檢查一次

    print(f"\n[Monitor] 開始監控未讀訊息...", flush=True)
    print(f"[Monitor] 停止方式：touch {STOP_FILE}", flush=True)

    # 多開 LINE 設定：BOXES 順序就是輪詢順序
    from line_locate import set_active_box, find_line_window
    import win32gui as _w32g
    BOXES = list(BOX_SOP.keys())   # 從 BOX_SOP 讀順序

    while is_before_stop_time(stop_time):
        if should_stop():
            break

        for box in BOXES:
            if should_stop():
                break
            try:
                # 切到指定 box 的 LINE 視窗 + 載入該 box 的 SOP
                set_active_box(box)
                box_sop = box_sop_cache.get(box)
                box_prompt = box_prompt_cache.get(box)
                if not box_sop or not box_prompt:
                    print(f"[Multi] box={box} 沒設定 SOP，跳過", flush=True)
                    continue

                line = find_line_window()
                if not line:
                    print(f"[Multi] 找不到 box={box} 的 LINE 視窗，跳過", flush=True)
                    continue
                hwnd = line[0]
                try:
                    _w32g.SetForegroundWindow(hwnd)
                except Exception as e:
                    print(f"[Multi] SetForegroundWindow {box} 失敗: {e}", flush=True)
                time.sleep(1.0)
                print(f"\n[Multi] === 開始處理 box={box} → {box_sop['name']} (hwnd={hwnd}) ===", flush=True)

                # Step 1-2: 處理未讀（每處理 1 個就重新偵測，避免 chat list 重排造成座標失效）
                # rename/分享後 LINE 會把對話置頂 → 後面的 y 座標會錯位 → 必須重 detect
                processed_count = 0
                MAX_PER_BOX = 10  # 每 box 單輪最多處理 10 個未讀（避免無限 loop）
                while processed_count < MAX_PER_BOX:
                    if should_stop():
                        break

                    regions, unread_list = find_unread_conversations(monitor)
                    if not unread_list:
                        if processed_count == 0:
                            print(f"[Multi] box={box} 無未讀，跳下一個", flush=True)
                        else:
                            print(f"[Multi] box={box} 本輪處理完畢（共 {processed_count} 個）", flush=True)
                        break

                    if processed_count == 0:
                        print(f"[Multi] box={box} 偵測到 {len(unread_list)} 個未讀對話", flush=True)

                    # 永遠處理最頂端未讀（重新 detect 後 unread_list[0] 是當下最新位置）
                    conv = unread_list[0]
                    regions = handle_one_customer(
                        conv, regions, box_prompt, box_sop, all_histories, monitor
                    )

                    if should_stop():
                        break

                    # 按 Esc 退出聊天室（不標已讀，對方再回覆會重新出現綠色徽章）
                    pyautogui.press("escape")
                    time.sleep(0.5)
                    processed_count += 1
                else:
                    # while 條件 processed_count < MAX_PER_BOX 失敗（達上限）
                    print(f"[Multi] box={box} 達單輪上限 {MAX_PER_BOX}，跳下一個", flush=True)

            except Exception as e:
                print(f"[ERR] box={box}: {e}", flush=True)
                import traceback
                traceback.print_exc()
                time.sleep(2)

        # 全部 box 跑完一輪，等待後再從頭
        for _ in range(POLL_INTERVAL):
            if should_stop():
                break
            time.sleep(1)

    reason = "收到停止信號" if _should_stop else "到達結束時間"
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 監控結束（{reason}）", flush=True)
    print(f"共處理 {len(all_histories)} 個客戶", flush=True)
    return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        print("用法：python line_auto_multi.py <監控時間HH:MM> [SOP設定檔.json]")
        sys.exit(0)

    stop = sys.argv[1]
    sop = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SOP
    main(stop, sop)
